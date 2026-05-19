import os
import time
import json
import platform
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes

# ============================================================
EMAIL = os.environ["ITLA_EMAIL"]
PASSWORD = os.environ["ITLA_PASSWORD"]
BOT_TOKEN = os.environ["BOT_TOKEN"]
CHAT_ID = int(os.environ["CHAT_ID"])
# ============================================================

INTERCEPTOR = """
window._calificaciones = null;
window._interceptorReady = true;
const originalFetch = window.fetch;
window.fetch = function(...args) {
    const request = args[0];
    const url = typeof request === 'string' ? request : request.url;
    console.log('[INTERCEPTOR] fetch:', url);
    const promise = originalFetch.apply(this, args);
    if (url.includes('GetCalificaciones') || url.includes('calificacion') || url.includes('Calificacion')) {
        promise.then(response => {
            response.clone().json().then(data => {
                console.log('[INTERCEPTOR] Datos capturados:', JSON.stringify(data).substring(0, 200));
                window._calificaciones = JSON.stringify(data);
            }).catch((e) => { console.log('[INTERCEPTOR] Error parsing JSON:', e); });
        }).catch((e) => { console.log('[INTERCEPTOR] Error en fetch:', e); });
    }
    return promise;
};
const origOpen = XMLHttpRequest.prototype.open;
const origSend = XMLHttpRequest.prototype.send;
XMLHttpRequest.prototype.open = function(method, url) {
    this._url = url;
    return origOpen.apply(this, arguments);
};
XMLHttpRequest.prototype.send = function() {
    if (this._url) {
        console.log('[INTERCEPTOR] XHR:', this._url);
        if (this._url.includes('GetCalificaciones') || this._url.includes('calificacion') || this._url.includes('Calificacion')) {
            this.addEventListener('load', function() {
                try {
                    console.log('[INTERCEPTOR] XHR datos capturados:', this.responseText.substring(0, 200));
                    window._calificaciones = this.responseText;
                } catch(e) { console.log('[INTERCEPTOR] XHR error:', e); }
            });
        }
    }
    return origSend.apply(this, arguments);
};
console.log('[INTERCEPTOR] Interceptor instalado correctamente');
"""

def get_calificaciones():
    print("🌐 Abriendo navegador...")
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")

    if platform.system() == "Linux":
        # Oracle / servidor Linux
        options.binary_location = "/usr/bin/chromium"
        service = webdriver.ChromeService(executable_path="/usr/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=options)
    else:
        # Windows / Mac — Selenium auto-detecta Chrome y chromedriver
        driver = webdriver.Chrome(options=options)
    data = None

    try:
        driver.get("https://campusvirtual.itla.edu.do/account/login")
        wait = WebDriverWait(driver, 20)
        usuario_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Usuario']")))
        time.sleep(1)
        usuario_input.send_keys(EMAIL)
        driver.find_element(By.XPATH, "//input[@placeholder='Contraseña']").send_keys(PASSWORD)
        login_btn = driver.find_element(By.XPATH, "//button[contains(text(), 'Iniciar Sesión')]")
        driver.execute_script("arguments[0].click();", login_btn)
        print("🔐 Login enviado, esperando carga...")
        time.sleep(5)

        # Verificar que el login fue exitoso
        current_url = driver.current_url
        print(f"📍 URL actual: {current_url}")

        # Inyectar interceptor DIRECTO en la página actual (fix principal)
        driver.execute_script(INTERCEPTOR)
        interceptor_ok = driver.execute_script("return window._interceptorReady === true;")
        print(f"🔧 Interceptor inyectado en página actual: {interceptor_ok}")

        # También inyectar para futuras navegaciones completas (respaldo)
        try:
            driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": INTERCEPTOR})
            print("🔧 Interceptor CDP también instalado (respaldo)")
        except Exception as e:
            print(f"⚠️ CDP falló (no crítico): {e}")

        # Buscar y clickear Perfil (JS click para evitar elementos superpuestos)
        perfil_link = wait.until(EC.presence_of_element_located((By.XPATH, "//a[contains(text(), 'Perfil')]")))
        print("🔗 Click en Perfil...")
        driver.execute_script("arguments[0].click();", perfil_link)

        # Esperar datos
        for i in range(40):
            time.sleep(1)
            result = driver.execute_script("return window._calificaciones;")

            if result is not None and result != "null":
                parsed = json.loads(result)
                if isinstance(parsed, list) and len(parsed) == 0:
                    print(f"   {i+1}s - API respondió lista vacía (sin calificaciones aún)")
                    data = parsed
                    break
                elif parsed:
                    print(f"✅ Datos capturados en {i+1}s!")
                    print(f"🔑 Campos disponibles: {list(parsed[0].keys()) if parsed else 'vacío'}")
                    data = parsed
                    break

            # Cada 10s verificar que el interceptor sigue vivo
            if i % 10 == 9:
                still_ready = driver.execute_script("return window._interceptorReady === true;")
                curr_url = driver.current_url
                print(f"   {i+1}s - Interceptor activo: {still_ready} | URL: {curr_url}")
                if not still_ready:
                    print("🔄 Re-inyectando interceptor...")
                    driver.execute_script(INTERCEPTOR)

            print(f"   Esperando... {i+1}s")

        if data is None:
            # Debug: capturar info de la página para diagnosticar
            page_title = driver.title
            curr_url = driver.current_url
            print(f"⚠️ Sin datos después de 40s | Título: {page_title} | URL: {curr_url}")
            # Capturar logs del browser
            try:
                logs = driver.get_log("browser")
                interceptor_logs = [l for l in logs if "INTERCEPTOR" in l.get("message", "")]
                if interceptor_logs:
                    print("📋 Logs del interceptor:")
                    for log in interceptor_logs[-10:]:
                        print(f"   {log['message']}")
                else:
                    print("📋 No hay logs del interceptor (ningún request interceptado)")
            except Exception:
                pass

    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        driver.quit()

    return data

async def notas(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.id != CHAT_ID:
        return

    await update.message.reply_text("⏳ Buscando tus calificaciones, espera ~30 segundos...")
    data = get_calificaciones()

    if data is not None:
        if len(data) == 0:
            await update.message.reply_text("📭 No hay calificaciones registradas aún para este cuatrimestre.")
        else:
            periodo = data[0].get("periodo", "")

            # ── Mensaje formateado en Telegram ──
            msg = f"📚 <b>Calificaciones — {periodo}</b>\n"
            msg += "━" * 28 + "\n\n"

            for i, m in enumerate(data, 1):
                clave = m.get("clave", "")
                asignatura = m.get("asignatura", "")
                asignaciones = m.get("asignaciones", "--")
                practicas = m.get("practicas", "--")
                p1 = m.get("primerParcial", "--")
                p2 = m.get("segundoParcial", "--")
                final = m.get("final", "--")
                total = m.get("total", 0)
                ausencias = m.get("ausencias", 0)
                permitidas = m.get("ausenciasPermitidas", 0)

                msg += f"<b>{i}. {asignatura}</b>\n"
                msg += f"   📌 Código: {clave}\n"
                msg += f"   📋 Asignaciones: {asignaciones} | Prácticas: {practicas}\n"
                msg += f"   📝 1er Parcial: {p1} | 2do Parcial: {p2}\n"
                msg += f"   🏁 Final: {final} | Total: <b>{total}</b>\n"
                msg += f"   🚪 Ausencias: {ausencias}/{permitidas}\n\n"

            msg += "━" * 28 + "\n"
            msg += f"✅ {len(data)} materias encontradas"

            await update.message.reply_text(msg, parse_mode="HTML")

            # ── Excel formateado ──
            rows = []
            for m in data:
                rows.append({
                    "Código": m.get("clave", ""),
                    "Asignatura": m.get("asignatura", ""),
                    "Asignaciones": m.get("asignaciones", "--"),
                    "Prácticas": m.get("practicas", "--"),
                    "1er Parcial": m.get("primerParcial", "--"),
                    "2do Parcial": m.get("segundoParcial", "--"),
                    "Final": m.get("final", "--"),
                    "Total": m.get("total", 0),
                    "Ausencias": m.get("ausencias", 0),
                    "Aus. Permitidas": m.get("ausenciasPermitidas", 0),
                })

            df = pd.DataFrame(rows)
            filepath = "calificaciones.xlsx"
            df.to_excel(filepath, index=False, sheet_name="Calificaciones")

            # Dar formato al Excel
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active

            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"),
            )

            # Aplicar estilos al header
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border

            # Aplicar estilos a las celdas y alternar colores
            light_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    if row % 2 == 0:
                        cell.fill = light_fill
                # Asignatura alineada a la izquierda
                ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center")

            # Auto-ajustar ancho de columnas
            for col in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = get_column_letter(col)
                for row in range(1, ws.max_row + 1):
                    val = str(ws.cell(row=row, column=col).value or "")
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = max_len + 4

            # Fila de header más alta
            ws.row_dimensions[1].height = 30

            wb.save(filepath)

            await update.message.reply_document(
                document=open(filepath, "rb"),
                filename=f"calificaciones_{periodo}.xlsx"
            )
    else:
        await update.message.reply_text("❌ No se pudieron obtener los datos. Revisa los logs en la consola.")

app = ApplicationBuilder().token(BOT_TOKEN).build()
app.add_handler(CommandHandler("notas", notas))
print("🤖 Bot activo. Escribe /notas en Telegram.")
app.run_polling()
